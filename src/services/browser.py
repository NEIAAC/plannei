from asyncio import subprocess
import random
from typing import Literal
from enum import Enum
from csv import DictReader, __version__ as csv_version

from PySide6.QtCore import QThread, Signal

from openpyxl import load_workbook, __version__ as openpyxl_version
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webelement import WebElement

from utils.logger import logger, LogLevel


class BrowserChoice(Enum):
    CHROME = "chrome"
    FIREFOX = "firefox"


USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36",
]

BASE_URL = "https://inforestudante.uc.pt"
LOGIN_URL = f"{BASE_URL}/nonio/security/login.do"
ENROLL_URL = f"{BASE_URL}/nonio/inscturmas/init.do"


class BrowserThread(QThread):
    outputSignal = Signal(str, str)

    def __init__(
        self,
        loginEmail: str,
        loginPassword: str,
        browserChoice: BrowserChoice,
        headless: bool,
        dryRun: bool,
        enrollmentIndex: int,
        tablePath: str,
    ):
        super().__init__()
        self.loginEmail = loginEmail
        self.loginPassword = loginPassword
        self.browserChoice = browserChoice
        self.headless = headless
        self.enrollmentIndex = enrollmentIndex
        self.tablePath = tablePath
        self.dryRun = dryRun

    def output(self, text: str, level: LogLevel = LogLevel.INFO):
        logger.log(level.value, text)
        self.outputSignal.emit(text, level.value)

    def readTable(self) -> tuple[list[dict[str, str]], list[str]]:
        if self.tablePath.endswith(".csv"):
            logger.info(
                f"Using native CSV {csv_version} module to read {self.tablePath}"
            )
            records = []
            with open(self.tablePath, mode="r", encoding="utf-8") as csvfile:
                reader = DictReader(csvfile)
                headers = reader.fieldnames
                if not headers:
                    raise ValueError("CSV file has no headers")
                logger.info(f"Headers found in CSV: {len(headers)}")
                for row in reader:
                    records.append(
                        {
                            key: (value if value else "")
                            for key, value in row.items()
                        }
                    )
        elif self.tablePath.endswith(".xlsx"):
            logger.info(
                f"Using openpyxl {openpyxl_version} to read {self.tablePath}"
            )
            workbook = load_workbook(filename=self.tablePath, data_only=True)
            logger.info(
                f"Loaded workbook with {len(workbook.sheetnames)} sheets"
            )
            sheet = workbook.active
            if not sheet:
                raise ValueError("Excel file has no sheets")
            logger.info(
                f"Active sheet: {sheet.title}, {sheet.max_row} rows, {sheet.max_column} columns"
            )
            headers = [str(cell.value) for cell in sheet[1] if cell.value]
            logger.info(f"Headers found: {len(headers)}")
            if not headers:
                raise ValueError("Excel file has no headers")
            records = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                logger.debug(f"Row data: {row}")
                record = {
                    headers[i]: (value if value else "")
                    for i, value in enumerate(row)
                    if i < len(headers)
                }
                records.append(record)
        else:
            raise ValueError("Unsupported file extension")

        return records, list(headers)

    def setupChromium(self) -> webdriver.Chrome:
        options = webdriver.ChromeOptions()
        options.add_argument("--disable-extensions")
        options.add_argument("--disable-gpu")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-setuid-sandbox")
        options.add_argument("--disable-popup-blocking")
        options.add_argument("--disable-infobars")
        options.add_argument("--disable-notifications")
        options.add_argument("--disable-default-apps")
        options.add_argument("--mute-audio")
        options.add_argument("--blink-settings=imagesEnabled=false")
        options.add_argument("--remote-allow-origins=*")
        options.add_argument(f"--user-agent={random.choice(USER_AGENTS)}")
        if self.headless:
            options.add_argument("--window-size=1920,1080")
            options.add_argument("--headless")
        else:
            options.add_argument("--start-maximized")
        options.add_experimental_option("excludeSwitches", ["enable-logging"])
        options.page_load_strategy = "eager"
        service = webdriver.ChromeService(log_output=subprocess.DEVNULL)
        driver = webdriver.Chrome(options=options, service=service)
        if not self.headless:
            driver.maximize_window()
        return driver

    def setupFirefox(self) -> webdriver.Firefox:
        options = webdriver.FirefoxOptions()
        options.set_preference("permissions.default.desktop-notification", 2)
        options.set_preference("permissions.default.image", 2)
        options.set_preference("dom.push.enabled", False)
        options.set_preference(
            "dom.webnotifications.serviceworker.enabled", False
        )
        options.set_preference("dom.webnotifications.enabled", False)
        options.set_preference(
            "general.useragent.override", random.choice(USER_AGENTS)
        )
        if self.headless:
            options.add_argument("--width=1920")
            options.add_argument("--height=1080")
            options.add_argument("--headless")
        options.page_load_strategy = "eager"
        service = webdriver.FirefoxService(log_output=subprocess.DEVNULL)
        driver = webdriver.Firefox(options=options, service=service)
        if not self.headless:
            driver.maximize_window()
        return driver

    def run(self):
        try:
            inputs = self.__dict__.copy()
            inputs.pop("loginEmail", False)
            inputs.pop("loginPassword", False)

            logger.info(
                f"Starting browser thread with input parameters: {inputs}"
            )

            self.output("...")

            try:
                rows_dict, headers = self.readTable()
            except Exception as error:
                self.output(f"Failed to load table: {error}", LogLevel.ERROR)
                return
            logger.info(
                f"Table loaded, found {len(rows_dict)} {len(rows_dict) == 1 and 'row' or 'rows'}"
            )
            logger.info(f"Table headers read: {headers}")
            logger.info(f"Table rows read: {rows_dict}")

            if not rows_dict or not headers:
                self.output("Given table is empty", LogLevel.ERROR)
                return
            if len(rows_dict) < 1:
                self.output(
                    "Given table must have at least 1 row IN addition to the headers",
                    LogLevel.ERROR,
                )
                return
            if len(headers) != 4:
                self.output(
                    "Given table must have exactly 4 columns",
                    LogLevel.ERROR,
                )
                return

            classes_dict: dict[
                str,
                dict[
                    Literal["href"]
                    | Literal["className"]
                    | Literal["PL"]
                    | Literal["TP"]
                    | Literal["T"],
                    str | list[str] | None,
                ],
            ] = {}
            for item in rows_dict:
                classId = item[headers[0]].replace(" ", "").split("#")[0]

                pl_array = (
                    item[headers[1]].replace(" ", "").split("#")
                    if item[headers[1]]
                    else None
                )
                tp_array = (
                    item[headers[2]].replace(" ", "").split("#")
                    if item[headers[2]]
                    else None
                )
                t_array = (
                    item[headers[3]].replace(" ", "").split("#")
                    if item[headers[3]]
                    else None
                )

                classes_dict[classId] = {
                    "href": None,
                    "PL": pl_array,
                    "TP": tp_array,
                    "T": t_array,
                }
            self.output(
                f"{'[DRY-RUN MODE] ' if self.dryRun else ''}Enrolling in the {len(classes_dict)} classes found in the schedule table",
            )

            self.output("Starting browser")
            try:
                if self.browserChoice == BrowserChoice.CHROME:
                    driver = self.setupChromium()
                elif self.browserChoice == BrowserChoice.FIREFOX:
                    driver = self.setupFirefox()
                else:
                    self.output(
                        "Invalid browser choice",
                        LogLevel.ERROR,
                    )
                    return
            except Exception as error:
                self.output(
                    f"No supported browser found, you need to have the selected browser installed on your system: {error}",
                    LogLevel.ERROR,
                )
                return

            self.output(f"{driver.name.capitalize()} initialized")
            self.output(
                f"Navigating to {LOGIN_URL.split('/')[-1].split('.')[0]}"
            )
            driver.get(LOGIN_URL)

            if driver.current_url != LOGIN_URL:
                self.output("Already logged in")
            else:
                username_input = driver.find_element(
                    By.CSS_SELECTOR, "input#username"
                )
                password_input = driver.find_element(
                    By.CSS_SELECTOR, "input#password1"
                )
                username_input.send_keys(self.loginEmail)
                password_input.send_keys(self.loginPassword)
                login_button = driver.find_element(
                    By.CSS_SELECTOR, "input[type='submit']"
                )
                login_button.click()
                if driver.current_url == LOGIN_URL:
                    self.output(
                        "Login failed, check your credentials and retry",
                        LogLevel.ERROR,
                    )
                    return
                self.output("Login successful")

            self.output(f"Now at {driver.current_url.split('/')[-2]}")
            self.output(f"Navigating to {ENROLL_URL.split('/')[-2]}")
            driver.get(ENROLL_URL)

            tableBody = driver.find_element(
                By.CSS_SELECTOR, "table.displaytable > tbody"
            )
            try:
                chosenEnrollment = tableBody.find_element(
                    By.CSS_SELECTOR,
                    f"tr:nth-of-type({self.enrollmentIndex})",
                )
            except Exception:
                self.output(
                    f"Failed to find enrollment with index {self.enrollmentIndex}",
                    LogLevel.ERROR,
                )
                return

            chosenEnrollmentText = chosenEnrollment.find_element(
                By.CSS_SELECTOR, "td:first-child"
            ).text
            chosenEnrollmentLink = chosenEnrollment.find_element(
                By.CSS_SELECTOR, "td:last-child > div > a"
            ).get_attribute("href")

            if not chosenEnrollmentLink:
                self.output("No enrollment link found", LogLevel.ERROR)
                return

            self.output(f"Proceeding to enrollment in {chosenEnrollmentText}")
            driver.get(chosenEnrollmentLink)

            # Parse every row in the courses table body and add the href for that courses enrollment page to the classes_dict
            tableBody = driver.find_element(
                By.CSS_SELECTOR, "table.displaytable > tbody"
            )
            for row in tableBody.find_elements(By.CSS_SELECTOR, "tr"):
                cells = row.find_elements(By.CSS_SELECTOR, "td")
                classId = cells[0].text.strip()
                className = cells[1].text.strip()
                href = (
                    cells[6]
                    .find_element(By.CSS_SELECTOR, "a")
                    .get_attribute("href")
                )
                if classId in classes_dict:
                    classes_dict[classId]["href"] = href
                    classes_dict[classId]["className"] = className

            picked_dict: dict[str, str | None] = {}
            for classId, classData in classes_dict.items():
                if not isinstance(classData["href"], str):
                    self.output(
                        f"No class found with ID {classId} from schedule table so it will be skipped, check for typos in ID",
                        LogLevel.WARNING,
                    )
                    continue
                self.output(f"Proceeding to {classData['className']} schedule")

                # Maybe this can be improved to use multiple tabs, open all course pages at once in
                # different tabs and fill the preferences in parallel for each one
                driver.get(classData["href"])

                if not self.dryRun:
                    try:
                        save_button = driver.find_element(By.ID, "botaoGravar")
                    except Exception:
                        self.output(
                            f"Schedule choice for {classData['className']} ({classId}) is not available yet, skipping",
                            LogLevel.WARNING,
                        )
                        continue
                else:
                    # Use back button instead of trying to find save in dry run mode
                    save_button = driver.find_element(By.ID, "botaoVoltar")

                classTypes = ("PL", "TP", "T")

                for classType in classTypes:
                    if classData[classType]:
                        classRows = driver.find_elements(
                            By.CSS_SELECTOR,
                            f'table[class="displaytable"]:has(input[type="checkbox"][alt="{classType}"]) > tbody > tr',
                        )
                        if not classRows:
                            self.output(
                                f"Class {classData['className']} ({classId}) had {classType} preferences but there are no {classType} schedules available",
                                LogLevel.WARNING,
                            )
                            continue
                        self.output(
                            f"Choosing {classType} schedule for class {classData['className']}"
                        )
                        preferences_dict: dict[str, WebElement] = {}
                        for row in classRows:
                            cells = row.find_elements(By.TAG_NAME, "td")

                            element = cells[0].get_attribute("innerHTML") or ""

                            classNumber = (
                                element.split("<sup>")[0]
                                .split(classType)[-1]
                                .strip()
                            )

                            if classNumber not in (classData[classType] or []):
                                continue

                            # Try to find preview checkbox instead of the real selection checkbox
                            # so dryRun works even if enrollments are open
                            if self.dryRun:
                                try:
                                    classCheckbox = cells[-2].find_element(
                                        By.TAG_NAME, "input"
                                    )
                                except Exception:
                                    classCheckbox = cells[-1].find_element(
                                        By.TAG_NAME, "input"
                                    )
                            else:
                                classCheckbox = cells[-1].find_element(
                                    By.TAG_NAME, "input"
                                )
                            if classCheckbox.get_attribute("disabled"):
                                self.output(
                                    f"{classType}{classNumber} is full (or mandatory) for class {classData['className']}, skipping",
                                    LogLevel.INFO,
                                )
                                continue
                            if classCheckbox.get_attribute("checked"):
                                self.output(
                                    f"Already enrolled in {classType}{classNumber} for class {classData['className']}, ",
                                    LogLevel.INFO,
                                )
                                continue
                            preferences_dict[classNumber] = classCheckbox
                        for classNumber in classData[classType] or []:
                            if classNumber in preferences_dict:
                                self.output(
                                    f"Enrolling in {classType}{classNumber} for class {classData['className']}"
                                )
                                preferences_dict[classNumber].click()
                                picked_dict[
                                    f"{classData['className']} {classType}"
                                ] = classNumber
                                break
                save_button.click()
            self.output(
                f"Enrollment completed for {len(picked_dict)} classes",
            )
            if picked_dict:
                self.output(
                    "Final choices were:",
                    LogLevel.SUCCESS,
                )
                for className, preference in picked_dict.items():
                    self.output(f"{className}{preference}", LogLevel.SUCCESS)

        except Exception as error:
            self.output(
                f"An unexpected error occurred: {f'{error}'.split(';')[0][9:]}",
                LogLevel.ERROR,
            )
